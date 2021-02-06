from selenium import webdriver
from selenium.webdriver.common.by import By
import valid
from valid import validation
from valid import announce
import openpyxl
import datetime

# 関数宣言
def count_elems(css_links_first):
    count = driver.find_elements_by_css_selector(css_links_first)
    elem_count = len(count) + 1
    return elem_count

def access_by_nth_child(kancho_name, css_links_first, elems, css_links_latter):
    for n in range(1, elems):
        css_links_perfect = css_links_first + ":nth-child(" + str(n) + css_links_latter
        print(kancho_name + " No." + str(n) + ": " + css_links_perfect)
        get_data(css_date, css_links_perfect, n)

def access_by_nth_of_type(kancho_name, css_links_first, elems, css_links_latter):
    for n in range(1, elems):
        css_links_perfect = css_links_first + ":nth-of-type(" + str(n) + css_links_latter
        print(kancho_name + " No." + str(n) + ": " + css_links_perfect)
        get_data(css_date, css_links_perfect, n)

def get_data(css_date, css_links_perfect, n):

    links = driver.find_elements(By.CSS_SELECTOR, css_links_perfect)
    for link in links:
        link_text = link.text
        link_url = link.get_attribute("href")
        data_list.append([kancho_name, date_text, link_text, link_url, n])
    return data_list

# URLリストの読み込み
wb = openpyxl.load_workbook("官庁新着情報URL.xlsx")
ws = wb["Sheet1"]

row_values_list = [] #URLとCSSセレクタを入れる配列

for row in ws.iter_rows(min_row=2): 
    if row[0].value is None:
        break

    value_list = []
    for c in row: # cという変数名に意味はない 各列の値をひとつの要素として、一行ぶんの値を配列にまとめる
        value_list.append(c.value)
    row_values_list.append(value_list) # 多次元配列として行ごとの値の配列を格納する

driver_path = "driver/chromedriver_mac87"
driver = webdriver.Chrome(executable_path=driver_path) # webdriverの作成
driver.implicitly_wait(5) # 要素が見つからなければ5秒待つ設定

data_list = [] # 読み取り結果を格納する

for value in row_values_list:
    kancho_name = str(value[0])
    kancho_url = str(value[1])
    css_date = str(value[2])
    css_links_first = str(value[3])
    css_links_latter = str(value[4])
    nth_of_type = bool(value[5])
    
    driver.get(kancho_url)

    date_elem = driver.find_element(By.CSS_SELECTOR, css_date)
    date_text = date_elem.text

    # 日付一致のバリデーション
    get_bool = validation(kancho_name, date_text)
    if get_bool == False :
        valid.announce(kancho_name, valid.today)  
        continue

    elems = count_elems(css_links_first)
    if nth_of_type == True:
        access_by_nth_of_type(kancho_name, css_links_first, elems, css_links_latter)
    else:
        access_by_nth_child(kancho_name, css_links_first, elems, css_links_latter)

driver.quit()

# 新しいブックへの記入と保存
wb_new = openpyxl.Workbook()
ws_new = wb_new.worksheets[0]

row_num = 1

for data in data_list :
    ws_new.cell(row_num, 1).value = data[0] #官庁名
    ws_new.cell(row_num, 2).value = data[1] #日付
    ws_new.cell(row_num, 3).value = data[2] #テキスト
    ws_new.cell(row_num, 4).value = data[3] #リンク
    ws_new.cell(row_num, 5).value = data[4] #n

    row_num += 1

wb_new.save(str(format(valid.today, '%Y%m%d')) + "官庁新着情報.xlsx")